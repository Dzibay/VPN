# -*- coding: utf-8 -*-
"""Analyze Wordstat export for SEO page relevance."""
import json
import re
from collections import defaultdict
from openpyxl import load_workbook

PATH = r"c:\Users\долбаеб\Downloads\wordstat_top_queries (22).xlsx"

# Competitor brands (informational/comparison pages possible but not primary landing)
COMPETITORS = {
    "ваня впн", "дядя ваня", "дядя впн", "дед впн", "радмин впн", "амнезия впн",
    "happ впн", "хапп впн", "outline", "wireguard", "proton", "nordvpn", "expressvpn",
    "surfshark", "windscribe", "tunnelbear", "cyberghost", "hotspot shield",
    "planet vpn", "planetvpn", "v2raytun", "v2ray", "shadowsocks", "psiphon",
    "betternet", "hide me", "hideme", "purevpn", "ipvanish", "mullvad",
    "турбо впн", "turbo vpn", "super vpn", "supervpn", "vpn master",
    "bot vpn", "бот впн", "telegram vpn", "телеграм впн",
}

# Irrelevant patterns
IRRELEVANT = [
    r"\bбез впн\b", r"\bбез vpn\b", r"не работает", r"не включается", r"не подключается",
    r"ошибка", r"сломал", r"удалить", r"отключить", r"выключить", r"отменить подписк",
    r"взлом", r"crack", r"кряк", r"торрент", r"порно", r"xxx", r"18\+",
    r"игра\b", r"игры\b", r"майнкрафт", r"roblox", r"roblox", r"фортнайт",
    r"сериал", r"фильм", r"кино\b", r"аниме", r"мульт",
    r"что такое впн\b", r"что такое vpn\b",  # too generic wiki intent
    r"впн это\b", r"vpn это\b",
    r"настройка роутер", r"openvpn server", r"создать впн", r"свой впн сервер",
    r"vpn server", r"корпоративн", r"site to site",
]

# Service/use-case keywords aligned with Podorozhnik VPN
SERVICES = {
    "youtube": ["youtube", "ютуб", "ютюб"],
    "instagram": ["instagram", "инстаграм", "инста"],
    "chatgpt": ["chatgpt", "chat gpt", "чатgpt", "чат gpt", "openai"],
    "gemini": ["gemini", "google ai", "bard"],
    "claude": ["claude", "клод"],
    "telegram": ["telegram", "телеграм", "телега"],
    "tiktok": ["tiktok", "тик ток", "тикток"],
    "discord": ["discord", "дискорд"],
    "whatsapp": ["whatsapp", "ватсап", "вотсап"],
    "twitter": ["twitter", "x.com", "твиттер"],
    "facebook": ["facebook", "фейсбук"],
    "spotify": ["spotify", "спотифай"],
    "netflix": ["netflix", "нетфликс"],
    "steam": ["steam", "стим"],
    "twitch": ["twitch", "твич"],
    "zoom": ["zoom", "зум"],
    "linkedin": ["linkedin", "линкедин"],
    "pinterest": ["pinterest"],
    "reddit": ["reddit", "реддит"],
    "snapchat": ["snapchat"],
    "roblox": ["roblox", "роблокс"],  # gaming but high volume - separate
}

PLATFORMS = {
    "android": ["android", "андроид", "андроиде"],
    "ios": ["iphone", "айфон", "ios", "ipad", "айпад"],
    "windows": ["windows", "виндовс", "винда", "пк", "pc", "компьютер"],
    "mac": ["mac", "macos", "макбук", "macbook"],
    "linux": ["linux", "линукс"],
    "router": ["роутер", "router", "маршрутизатор"],
    "tv": ["smart tv", "телевизор", "android tv", "apple tv"],
    "browser_extension": ["расширение", "extension", "chrome", "яндекс браузер", "opera"],
}

INTENTS = {
    "free_trial": ["бесплатн", "пробн", "trial", "free", "без оплат", "без карты", "тест"],
    "download": ["скачать", "download", "установ", "install", "apk", "app"],
    "best_compare": ["лучший", "лучшие", "топ", "рейтинг", "какой", "какой выбрать", "сравнен"],
    "buy_price": ["купить", "цена", "стоимость", "подписк", "тариф", "оплат"],
    "russia": ["росси", "рф", "в россии", "для россии", "2026", "2025"],
    "split_tunnel": ["split", "сплит", "раздельн", "маршрутизац", "умный", "smart"],
    "vless": ["vless", "xray", "v2ray", "wireguard", "outline", "shadowsocks"],
    "unblock": ["обход", "разблок", "заблок", "блокиров", "доступ", "работает"],
    "speed": ["быстр", "скорост", "стабильн", "без потери"],
    "privacy": ["без лог", "no log", "аноним", "приват", "конфиденц", "безопас"],
    "for_service": ["для "],  # handled separately
}

GEO = ["герман", "сша", "usa", "европ", "китай", "япон", "турц", "казахстан", "беларус"]

def norm(s):
    return str(s or "").strip().lower()

def is_competitor(q):
    for c in COMPETITORS:
        if c in q:
            return True
    return False

def is_irrelevant(q):
    for pat in IRRELEVANT:
        if re.search(pat, q):
            return True
    return False

def detect_service(q):
    for name, kws in SERVICES.items():
        if any(k in q for k in kws):
            return name
    return None

def detect_platform(q):
    found = []
    for name, kws in PLATFORMS.items():
        if any(k in q for k in kws):
            found.append(name)
    return found

def detect_intents(q):
    found = set()
    for name, kws in INTENTS.items():
        if name == "for_service":
            continue
        if any(k in q for k in kws):
            found.add(name)
    return found

def score_relevance(q, count):
    """Higher = better SEO landing page candidate."""
    if is_competitor(q) or is_irrelevant(q):
        return -1, "exclude"

    score = 0
    reasons = []

    if "впн" in q or "vpn" in q:
        score += 2
        reasons.append("vpn-intent")

    svc = detect_service(q)
    if svc:
        score += 5
        reasons.append(f"service:{svc}")

    plats = detect_platform(q)
    if plats:
        score += 3
        reasons.append(f"platform:{','.join(plats)}")

    intents = detect_intents(q)
    commercial = {"download", "best_compare", "buy_price", "free_trial", "russia"}
    if intents & commercial:
        score += 4
        reasons.append(f"intent:{','.join(sorted(intents & commercial))}")

    if intents & {"split_tunnel", "vless", "unblock", "speed", "privacy"}:
        score += 3
        reasons.append("brand-fit")

    if any(g in q for g in GEO):
        score += 1
        reasons.append("geo")

    # Pure head terms
    if q in ("впн", "vpn", "лучший впн", "бесплатный впн", "скачать впн"):
        score += 3
        reasons.append("head-term")

    # Penalize very long tail with low specificity
    if len(q.split()) > 6 and score < 6:
        score -= 1
        reasons.append("long-tail")

    if score >= 7:
        tier = "A"
    elif score >= 5:
        tier = "B"
    elif score >= 3:
        tier = "C"
    else:
        tier = "D"

    return score, tier, reasons

def main():
    wb = load_workbook(PATH, read_only=True, data_only=True)
    ws = wb["Data"]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True)):
        q, count = row[0], row[1]
        if not q:
            continue
        q = norm(q)
        try:
            count = int(count or 0)
        except (TypeError, ValueError):
            count = 0
        rows.append((q, count))

    wb.close()
    print(f"Total queries: {len(rows)}")

    categorized = defaultdict(list)
    page_clusters = defaultdict(list)

    for q, count in rows:
        result = score_relevance(q, count)
        if result[0] == -1:
            categorized["exclude"].append((q, count, result[1]))
            continue
        score, tier, reasons = result
        item = (q, count, score, tier, reasons)
        categorized[tier].append(item)

        # Cluster into suggested pages
        svc = detect_service(q)
        plats = detect_platform(q)
        if svc and plats:
            key = f"/vpn-dlya-{svc}/{'-'.join(plats[:1])}"
        elif svc:
            key = f"/vpn-dlya-{svc}"
        elif plats and any(x in q for x in ["скачать", "бесплат", "лучш", "впн"]):
            key = f"/vpn-{'-'.join(plats[:1])}"
        elif "бесплатн" in q or "пробн" in q:
            key = "/besplatnyj-vpn"
        elif "лучш" in q or "какой" in q or "топ" in q:
            key = "/luchshij-vpn"
        elif "росси" in q or "рф" in q or "2026" in q:
            key = "/vpn-v-rossii"
        elif any(k in q for k in ["split", "сплит", "маршрутизац", "умный"]):
            key = "/split-tunneling-vpn"
        elif any(k in q for k in ["chatgpt", "gemini", "claude", "youtube", "ютуб"]):
            key = "/vpn-dlya-servisov"
        elif score >= 5:
            key = "/vpn-obshie"
        else:
            key = None

        if key and tier in ("A", "B"):
            page_clusters[key].append((q, count, score))

    # Output summary
    for tier in ["A", "B", "C", "D", "exclude"]:
        items = sorted(categorized.get(tier, []), key=lambda x: (-x[1] if tier != "exclude" else -x[1]))
        print(f"\n=== TIER {tier}: {len(items)} queries ===")
        for item in items[:40]:
            if tier == "exclude":
                print(f"  {item[1]:>10} | {item[0]} [{item[2]}]")
            else:
                print(f"  {item[1]:>10} | score={item[2]} | {item[0]}")

    print("\n=== SUGGESTED SEO PAGE CLUSTERS (A+B) ===")
    cluster_summary = []
    for page, items in sorted(page_clusters.items(), key=lambda x: -sum(i[1] for i in x[1])):
        total_vol = sum(i[1] for i in items)
        top = sorted(items, key=lambda x: -x[1])[:8]
        cluster_summary.append({
            "page": page,
            "total_volume": total_vol,
            "query_count": len(items),
            "top_queries": [(q, c) for q, c, _ in top],
        })
        print(f"\n{page} — {len(items)} queries, ~{total_vol:,} shows/mo")
        for q, c, sc in top:
            print(f"    {c:>10} | {q}")

    # Export JSON for reference
    out = {
        "total": len(rows),
        "tiers": {t: len(categorized.get(t, [])) for t in ["A", "B", "C", "D", "exclude"]},
        "clusters": cluster_summary,
        "tier_a": [(q, c, s, r) for q, c, s, t, r in categorized.get("A", [])],
        "tier_b": [(q, c, s, r) for q, c, s, t, r in categorized.get("B", [])],
    }
    out_path = r"c:\Users\долбаеб\Desktop\VPN\scripts\wordstat_analysis.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"\nSaved: {out_path}")

if __name__ == "__main__":
    main()
